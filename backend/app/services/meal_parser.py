"""주간 식단표 엑셀 파서 — 실제 '행복한 주간식단표(7.27~8.2)' 양식 기준.

구조: '일자' 행에 7일 날짜 → 아래로 상쾌한아침/간식/행복한점심/간식/편안한저녁 섹션.
섹션 라벨은 B열(셀 병합)에 세로로 들어있고, 이어지는 행은 라벨이 비어 있다.
맨 아래 '*', '◈' 줄은 안내 문구(원산지 등)로 notes에 담는다.
"""
from __future__ import annotations
import io
import re
from datetime import date, datetime, timedelta
from typing import Dict, List, Optional

MEAL_KEYS = ["아침", "간식(오전)", "점심", "간식(오후)", "저녁"]


def _as_date(v) -> Optional[date]:
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    try:
        f = float(v)
    except (TypeError, ValueError):
        return None
    if f > 40000:  # 엑셀 시리얼
        return date(1899, 12, 30) + timedelta(days=int(f))
    return None


def parse_meal_xlsx(data: bytes) -> Dict:
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    grid = [list(r) for r in ws.iter_rows(values_only=True)]

    # '일자' 행 찾기 — 날짜가 5개 이상 있는 행
    date_cols: List[int] = []
    dates: List[date] = []
    header_ri = -1
    for ri, row in enumerate(grid[:12]):
        found = [(ci, _as_date(v)) for ci, v in enumerate(row or []) if _as_date(v)]
        if len(found) >= 5:
            header_ri = ri
            date_cols = [ci for ci, _ in found]
            dates = [d for _, d in found]
            break
    if header_ri < 0:
        raise ValueError("날짜 행(일자)을 찾지 못했습니다 — 주간식단표 양식인지 확인해주세요.")

    # 섹션 스캔 — 라벨 열은 날짜 열 바로 왼쪽
    label_ci = max(0, min(date_cols) - 1)
    days: Dict[str, Dict[str, List[str]]] = {d.isoformat(): {k: [] for k in MEAL_KEYS} for d in dates}
    notes: List[str] = []
    cur: Optional[str] = None
    last_main: Optional[str] = None
    for ri in range(header_ri + 1, len(grid)):
        row = grid[ri] or []
        label_raw = row[label_ci] if label_ci < len(row) else None
        label = re.sub(r"\s+", "", str(label_raw)) if label_raw is not None else ""
        if label.startswith("*") or label.startswith("◈"):
            notes.append(str(label_raw).strip())
            cur = None
            continue
        if label:
            if "요일" in label:
                continue
            if "아침" in label:
                cur = last_main = "아침"
            elif "점심" in label:
                cur = last_main = "점심"
            elif "저녁" in label:
                cur = last_main = "저녁"
            elif "간식" in label:
                cur = "간식(오전)" if last_main == "아침" else "간식(오후)" if last_main == "점심" else "간식(오후)"
            else:
                cur = None
                continue
        if not cur:
            continue
        for ci, d in zip(date_cols, dates):
            v = row[ci] if ci < len(row) else None
            if v is None:
                continue
            for part in str(v).split("\n"):
                t = part.strip()
                if t:
                    days[d.isoformat()][cur].append(t)

    if not any(any(m for m in v.values()) for v in days.values()):
        raise ValueError("식단 내용을 읽지 못했습니다 — 시트 구조를 확인해주세요.")
    start = min(dates).isoformat()
    return {"start": start, "end": max(dates).isoformat(), "days": days, "notes": notes}
