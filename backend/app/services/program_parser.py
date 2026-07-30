"""프로그램 엑셀 파서 — 실제 '행복한_프로그램일정표/분류' 양식 기준.

일정표: 월 시트('26.7월') = 달력형. 날짜 행(숫자) 다음 두 행이 오전·오후 프로그램.
        셀 안 '[인지A]\n색칠공부' → group 태그 + 제목. 태그 없는 줄은 공통 프로그램.
분류표: 날짜 시트('26.07.24') = 인지/여가/신체 A·B·C 그룹별 명단 + 종교 명단.
"""
from __future__ import annotations
import io
import re
from typing import Any, Dict, List, Optional

GROUP_TAG = re.compile(r"\[(인지|여가|신체|신체기능)\s*([ABC])\]")
ANY_TAG = re.compile(r"\[([^\]]+)\]")
EDU_RGB = "FF00FF"  # 마젠타(분홍) 글씨 = 의무교육 항목 (노인인권·치매예방·감염예방 등)


def _is_edu_rgb(rgb) -> bool:
    return isinstance(rgb, str) and rgb.upper().endswith(EDU_RGB)


def _collect_edu(data: bytes, sheet_name: str) -> Dict:
    """분홍(마젠타) 리치텍스트 런/셀을 (행,열) 0기준 좌표별 텍스트 목록으로 수집."""
    import openpyxl
    out: Dict = {}
    try:
        from openpyxl.cell.rich_text import CellRichText
        wb = openpyxl.load_workbook(io.BytesIO(data), rich_text=True)
    except (TypeError, ImportError):   # 구버전 openpyxl — 색 구분 없이 진행
        return out
    ws = wb[sheet_name]
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 40), max_col=7):
        for c in row:
            v = c.value
            if v is None:
                continue
            key = (c.row - 1, c.column - 1)
            if isinstance(v, CellRichText):
                for part in v:
                    font = getattr(part, "font", None)
                    if font is not None and _is_edu_rgb(getattr(getattr(font, "color", None), "rgb", None)):
                        t = str(part).strip()
                        if t:
                            out.setdefault(key, []).append(t)
            elif isinstance(v, str) and v.strip():
                if _is_edu_rgb(getattr(getattr(c.font, "color", None), "rgb", None)):
                    out.setdefault(key, []).append(v.strip())
    return out


def _grid(ws) -> List[List[Any]]:
    return [list(row) for row in ws.iter_rows(values_only=True)]


def _month_from_sheet(name: str) -> Optional[str]:
    m = re.match(r"(\d{2})\.(\d{1,2})월", name.strip())
    if not m:
        return None
    return f"20{m.group(1)}-{int(m.group(2)):02d}"


def _parse_cell_entries(text: str, slot: str) -> List[Dict]:
    """'[인지A]\n색칠공부♥\n[기독교]' → [{slot, group:'인지A', title:'색칠공부♥'}, {slot, group:'기독교', title:''}]"""
    out: List[Dict] = []
    cur_group: Optional[str] = None
    cur_lines: List[str] = []

    def flush():
        nonlocal cur_group, cur_lines
        title = " ".join(l.strip() for l in cur_lines if l.strip())
        kind = None
        if "♥" in title:               # ♥ = 자체 프로그램 (외부강사 아님)
            kind = "자체"
            title = title.replace("♥", "").strip()
        if cur_group or title:
            out.append({"slot": slot, "group": cur_group, "title": title,
                        **({"kind": kind} if kind else {})})
        cur_group, cur_lines = None, []

    for line in str(text).split("\n"):
        line = line.strip()
        if not line:
            continue
        gm = GROUP_TAG.search(line)
        am = ANY_TAG.search(line)
        if gm:
            flush()
            cat = "신체" if gm.group(1).startswith("신체") else gm.group(1)
            cur_group = f"{cat}{gm.group(2)}"
            rest = ANY_TAG.sub("", line).strip()
            if rest:
                cur_lines.append(rest)
        elif am:      # [기독교], [자원봉사], [사회적응] 같은 일반 태그
            flush()
            cur_group = am.group(1).strip()
            rest = ANY_TAG.sub("", line).strip()
            if rest:
                cur_lines.append(rest)
        else:
            cur_lines.append(line)
    flush()
    return out


def list_schedule_months(data: bytes) -> List[str]:
    """엑셀 안의 월 시트('26.8월' 등)를 'YYYY-MM' 목록(최신순)으로."""
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True)
    out = sorted({m for n in wb.sheetnames if (m := _month_from_sheet(n))}, reverse=True)
    return out


def parse_schedule_xlsx(data: bytes, month: Optional[str] = None) -> Dict:
    """month('YYYY-MM')를 주면 그 달 시트를, 없으면 가장 최근 월 시트를 파싱."""
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    months = [(n, _month_from_sheet(n)) for n in wb.sheetnames]
    months = [(n, m) for n, m in months if m]
    if not months:
        raise ValueError("월 시트(예: '26.7월')를 찾지 못했습니다.")
    if month:
        picked = [(n, m) for n, m in months if m == month]
        if not picked:
            raise ValueError(f"{int(month[5:7])}월 시트가 엑셀에 없습니다.")
        sheet_name, month = picked[-1]
    else:
        sheet_name, month = max(months, key=lambda x: x[1])
    ws = wb[sheet_name]
    grid = _grid(ws)
    edu_map = _collect_edu(data, sheet_name)   # 분홍 글씨 = 교육

    days: Dict[str, List[Dict]] = {}
    y, mo = int(month[:4]), int(month[5:7])

    # 날짜 판정 — 숫자(1~31), 엑셀 시리얼, datetime 모두 지원.
    # 반환: 이 달의 일(int) / 다른 달 날짜(-1, 행 판정에는 포함) / 날짜 아님(None)
    from datetime import date, datetime as _dt, timedelta
    def day_of(v):
        if isinstance(v, _dt) or isinstance(v, date):
            return v.day if (v.year, v.month) == (y, mo) else -1
        try:
            f = float(v)
        except (TypeError, ValueError):
            return None
        if 1 <= f <= 31 and f == int(f):
            return int(f)
        if f > 40000:  # 엑셀 시리얼
            d = date(1899, 12, 30) + timedelta(days=int(f))
            return d.day if (d.year, d.month) == (y, mo) else -1
        return None

    # 날짜 행 = 비어 있지 않은 A~G 칸이 '전부' 날짜형 (첫 주는 전달 datetime + 이번 달 1일,
    # 마지막 주는 30·31 두 칸뿐일 수 있어 개수 기준은 위험하다)
    date_rows = []
    for ri, row in enumerate(grid[:40]):
        vals = [(row[c] if c < len(row) else None) for c in range(7)]
        marks = [day_of(v) if (v is not None and str(v).strip()) else None for v in vals]
        filled = [m for v, m in zip(vals, marks) if v is not None and str(v).strip()]
        if filled and all(m is not None for m in filled):
            date_rows.append((ri, [m if (m is not None and m > 0) else None for m in marks]))

    for bi, (ri, nums) in enumerate(date_rows):
        next_ri = date_rows[bi + 1][0] if bi + 1 < len(date_rows) else min(ri + 3, len(grid))
        for col in range(7):
            day = nums[col]
            if day is None:
                continue
            entries: List[Dict] = []
            for off, slot in ((1, "오전"), (2, "오후")):
                r2 = ri + off
                if r2 >= next_ri or r2 >= len(grid):
                    break
                cell = grid[r2][col] if col < len(grid[r2]) else None
                if cell is None or not str(cell).strip():
                    continue
                text = str(cell)
                for frag in edu_map.get((r2, col), []):   # 분홍(교육) 부분은 따로 뗀다
                    text = text.replace(frag, "")
                entries.extend(_parse_cell_entries(text, slot))
                for frag in edu_map.get((r2, col), []):
                    entries.append({"slot": slot, "group": None,
                                    "title": " ".join(frag.split()), "kind": "교육"})
            if entries:
                days[str(day)] = entries
    if not days:
        raise ValueError(f"'{sheet_name}' 시트에서 프로그램을 읽지 못했습니다.")

    # 오른쪽 규칙 메모 — ▶ 로 시작하는 운영 규칙, ♥(자체 프로그램)·외부강사 설명 등
    notes: List[str] = []
    seen = set()
    for row in grid[:30]:
        for ci in range(7, min(len(row or []), 22)):
            v = row[ci]
            if v is None:
                continue
            for line in str(v).split("\n"):
                t = line.strip()
                if not t or t in seen:
                    continue
                if t.startswith("▶") or t.startswith("*") or "외부강사" in t or "♥" in t:
                    notes.append(t)
                    seen.add(t)
    return {"month": month, "days": days, "sheet": sheet_name, "notes": notes}


def parse_groups_xlsx(data: bytes) -> Dict:
    """가장 최근 날짜 시트에서 그룹별 명단을 읽는다."""
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    dated = []
    for n in wb.sheetnames:
        m = re.match(r"(\d{2})\.(\d{2})\.(\d{2})", n.strip())
        if m:
            dated.append((n, f"20{m.group(1)}-{m.group(2)}-{m.group(3)}"))
    if not dated:
        raise ValueError("날짜 시트(예: '26.07.24')를 찾지 못했습니다.")
    sheet_name, based_on = max(dated, key=lambda x: x[1])
    ws = wb[sheet_name]
    grid = _grid(ws)

    HEADER = re.compile(r"(인지|여가|신체기능|신체)\s*([ABC])\s*그룹")
    groups: List[Dict] = []
    religion: List[Dict] = []

    def names_from(v) -> List[str]:
        return [w for w in re.split(r"[\s\n]+", str(v or "").strip()) if w and w != "None"]

    for ri, row in enumerate(grid):
        for ci, cell in enumerate(row or []):
            if cell is None:
                continue
            hm = HEADER.search(str(cell))
            if hm:
                cat = "신체" if hm.group(1).startswith("신체") else hm.group(1)
                members: List[str] = []
                by_floor: Dict[str, List[str]] = {}
                # 헤더 아래 최대 6행에서 '층' 행의 대상자(다음 열) 수집 — 층별로도 보존
                for r2 in range(ri + 1, min(ri + 7, len(grid))):
                    label = str(grid[r2][ci] or "").strip() if ci < len(grid[r2]) else ""
                    if HEADER.search(label):
                        break
                    fm = re.match(r"(\d층)", label)
                    if fm:
                        val = grid[r2][ci + 1] if ci + 1 < len(grid[r2]) else None
                        names = names_from(val)
                        if names:
                            by_floor.setdefault(fm.group(1), []).extend(names)
                        members.extend(names)
                groups.append({"category": cat, "grade": hm.group(2),
                               "members": members, "members_by_floor": by_floor})
            elif "종" in str(cell) and "교" in str(cell) and "활" in str(cell):
                for r2 in range(ri + 1, min(ri + 12, len(grid))):
                    label = str(grid[r2][ci] or "").strip() if ci < len(grid[r2]) else ""
                    if label in ("기독교", "천주교", "불교", "유교", "무교", "원불교"):
                        val = grid[r2][ci + 1] if ci + 1 < len(grid[r2]) else None
                        religion.append({"name": label, "members": names_from(val)})
    if not groups:
        raise ValueError(f"'{sheet_name}' 시트에서 그룹을 읽지 못했습니다.")
    return {"based_on": based_on, "groups": groups, "religion": religion, "sheet": sheet_name}
