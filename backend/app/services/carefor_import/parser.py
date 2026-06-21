"""
케어포 엑셀 파서
xlsx → openpyxl / xls → xlrd / csv → csv
절대 xls를 openpyxl로 읽지 않는다.
"""
import csv
import io
import re
from datetime import datetime, date, time
from typing import List, Dict, Any


def parse_file(file_bytes: bytes, filename: str) -> List[Dict[str, Any]]:
    """파일을 읽어 행 딕셔너리 리스트로 반환"""
    ext = filename.rsplit('.', 1)[-1].lower()

    if ext == 'xlsx':
        return _parse_xlsx(file_bytes)
    elif ext == 'xls':
        return _parse_xls(file_bytes)
    elif ext == 'csv':
        return _parse_csv(file_bytes)
    else:
        raise ValueError(f"지원하지 않는 파일 형식: {ext}")


def _parse_xlsx(file_bytes: bytes) -> List[Dict[str, Any]]:
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError("pip install openpyxl")

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    rows = []
    for sheet in wb.worksheets[:3]:
        headers = []
        for i, row in enumerate(sheet.iter_rows(values_only=True)):
            # 빈 행 스킵
            if all(c is None or str(c).strip() == '' for c in row):
                continue
            if not headers:
                headers = [str(c).strip() if c is not None else f"col_{j}"
                           for j, c in enumerate(row)]
                continue
            record = {}
            for j, val in enumerate(row):
                key = headers[j] if j < len(headers) else f"col_{j}"
                record[key] = val
            if any(v is not None and str(v).strip() for v in record.values()):
                rows.append(record)
    return rows


def _parse_xls(file_bytes: bytes) -> List[Dict[str, Any]]:
    try:
        import xlrd
    except ImportError:
        raise RuntimeError("pip install xlrd")

    book = xlrd.open_workbook(file_contents=file_bytes)
    rows = []
    for sheet in book.sheets()[:3]:
        headers = []
        for i in range(sheet.nrows):
            row_vals = sheet.row_values(i)
            if all(str(v).strip() == '' for v in row_vals):
                continue
            if not headers:
                headers = [str(v).strip() if str(v).strip() else f"col_{j}"
                           for j, v in enumerate(row_vals)]
                continue
            record = {}
            for j, val in enumerate(row_vals):
                key = headers[j] if j < len(headers) else f"col_{j}"
                # xlrd 날짜 serial 처리
                if sheet.cell_type(i, j) == 3:  # xlrd.XL_CELL_DATE
                    try:
                        import xlrd as _xlrd
                        dt = _xlrd.xldate_as_datetime(val, book.datemode)
                        val = dt.strftime('%Y-%m-%d')
                    except Exception:
                        pass
                record[key] = val if val != '' else None
            if any(v is not None and str(v).strip() for v in record.values()):
                rows.append(record)
    return rows


def _parse_csv(file_bytes: bytes) -> List[Dict[str, Any]]:
    text = file_bytes.decode('utf-8-sig', errors='replace')
    reader = csv.DictReader(text.splitlines())
    return [
        {k: (v if v else None) for k, v in row.items()}
        for row in reader
        if any(v for v in row.values())
    ]


# ══════════════════════════════════════════════════════════════════════════════
# 외박/외출 전용 파서 — 5행 헤더, 6행부터 데이터 (OpenAI 불필요, 결정적 파싱)
# ══════════════════════════════════════════════════════════════════════════════

LEAVE_HEADER_ROW = 5   # 1-based: 5행이 헤더


def _clean_header(h: Any) -> str:
    """헤더명 정규화
    - 줄바꿈/공백 제거: '외박\n일수' → '외박일수'
    - 괄호 설명 제거: '보호자\n(관계, 연락처)' → '보호자'
    """
    if h is None:
        return ''
    s = str(h)
    s = re.sub(r'\([^)]*\)', '', s)   # (관계, 연락처) 등 괄호 묶음 제거
    s = re.sub(r'\s+', '', s)          # 줄바꿈·공백 모두 제거
    return s.strip()


def _cell_value(v: Any) -> Any:
    """셀 값 정규화 — 날짜/시간 객체는 문자열로, 빈 칸은 None."""
    if v is None:
        return None
    if isinstance(v, datetime):
        if v.hour == 0 and v.minute == 0 and v.second == 0:
            return v.strftime('%Y-%m-%d')
        return v.strftime('%Y-%m-%d %H:%M')
    if isinstance(v, date):
        return v.strftime('%Y-%m-%d')
    if isinstance(v, time):
        return v.strftime('%H:%M')
    s = str(v).strip()
    return s if s else None


def _is_total_row(values: List[Any]) -> bool:
    """합계/공백 표시 행 판별 ('* 전체', '합계', '소계', '총계' 등)"""
    cells = [str(v).strip() for v in values if v is not None and str(v).strip() != '']
    if not cells:
        return False
    first = cells[0].replace(' ', '')
    if first.startswith('*'):
        return True
    if first in ('합계', '소계', '총계', '계', '전체', '총원', '현원'):
        return True
    for v in values:
        if v is not None and '*전체' in str(v).replace(' ', ''):
            return True
    return False


def _rows_from_grid(grid: List[List[Any]], header_row: int = LEAVE_HEADER_ROW) -> List[Dict[str, Any]]:
    hidx = header_row - 1
    if len(grid) <= hidx:
        return []

    raw_headers = grid[hidx]
    headers = []
    for j, h in enumerate(raw_headers):
        name = _clean_header(h)
        headers.append(name if name else f"col_{j}")

    out: List[Dict[str, Any]] = []
    for values in grid[hidx + 1:]:
        if not values:
            continue
        if all(v is None or str(v).strip() == '' for v in values):
            continue
        if _is_total_row(values):
            continue
        record: Dict[str, Any] = {}
        for j, key in enumerate(headers):
            record[key] = _cell_value(values[j] if j < len(values) else None)
        if any(v is not None for v in record.values()):
            out.append(record)
    return out


def _xlsx_grid(file_bytes: bytes) -> List[List[Any]]:
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.worksheets[0]
    return [list(row) for row in ws.iter_rows(values_only=True)]


def _xls_grid(file_bytes: bytes) -> List[List[Any]]:
    import xlrd
    book = xlrd.open_workbook(file_contents=file_bytes)
    sheet = book.sheets()[0]
    grid = []
    for i in range(sheet.nrows):
        vals = sheet.row_values(i)
        out = []
        for j, v in enumerate(vals):
            if sheet.cell_type(i, j) == 3:   # XL_CELL_DATE
                try:
                    v = xlrd.xldate_as_datetime(v, book.datemode).strftime('%Y-%m-%d')
                except Exception:
                    pass
            out.append(v)
        grid.append(out)
    return grid


def _csv_grid(file_bytes: bytes) -> List[List[Any]]:
    text = file_bytes.decode('utf-8-sig', errors='replace')
    return [row for row in csv.reader(text.splitlines())]


def parse_5row_file(file_bytes: bytes, filename: str,
                    header_row: int = LEAVE_HEADER_ROW) -> List[Dict[str, Any]]:
    """5행 헤더 기준 범용 파서 (외박/외출·수급자 공통).
    규칙: 5행 헤더, 6행부터 데이터, 헤더명 매핑, 줄바꿈 헤더 합침,
          빈 칸 None, 합계/공백 행 제외.
    """
    ext = filename.rsplit('.', 1)[-1].lower()
    if ext == 'xlsx':
        grid = _xlsx_grid(file_bytes)
    elif ext == 'xls':
        grid = _xls_grid(file_bytes)
    elif ext == 'csv':
        grid = _csv_grid(file_bytes)
    else:
        raise ValueError(f"지원하지 않는 파일 형식: {ext}")
    return _rows_from_grid(grid, header_row)


# 하위 호환 별칭
parse_leave_file = parse_5row_file
