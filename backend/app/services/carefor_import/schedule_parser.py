"""
근무표 Rule-Based Parser
지원 양식:
  헤더 1행: 직종 | 성명 | 1 | 2 | ... | 31 | 갯수 | 시간 | ...
  헤더 2행: (요일) 금 | 토 | 일 | ...
  데이터행: 직종 | 직원명(또는 조명) | D/N/휴/... | ...

OpenAI는 rule-based 파싱이 실패한 경우에만 fallback으로 사용
"""
import re
import io
import logging
from typing import List, Dict, Any, Optional, Tuple

logger = logging.getLogger(__name__)

# ── 날짜 컬럼 이후 무시할 비-근무 컬럼 키워드 ─────────────────────────────────
IGNORE_COL_KEYWORDS = {"갯수", "시간", "대휴", "일수", "비고", "총시간", "계", "합계", "비고란"}

# ── Shift 맵 ──────────────────────────────────────────────────────────────────
SHIFT_MAP: Dict[str, Dict] = {
    "D":   {"label": "주간",   "start": "09:00", "end": "18:00", "is_working": True},
    "N":   {"label": "야간",   "start": "18:00", "end": "09:00", "is_working": True},
    "AD":  {"label": "추가주간", "start": "09:00", "end": "18:00", "is_working": True},
    "PD":  {"label": "추가주간", "start": "09:00", "end": "18:00", "is_working": True},
    "E":   {"label": "이브닝", "start": "13:00", "end": "22:00", "is_working": True},
    "주":  {"label": "주간",   "start": "09:00", "end": "18:00", "is_working": True},
    "야":  {"label": "야간",   "start": "18:00", "end": "09:00", "is_working": True},
    "주간": {"label": "주간",  "start": "09:00", "end": "18:00", "is_working": True},
    "야간": {"label": "야간",  "start": "18:00", "end": "09:00", "is_working": True},
    "휴":  {"label": "휴무",   "start": None, "end": None, "is_working": False},
    "休":  {"label": "휴무",   "start": None, "end": None, "is_working": False},
    "휴무": {"label": "휴무",  "start": None, "end": None, "is_working": False},
    "대휴": {"label": "대체휴무","start": None, "end": None, "is_working": False},
    "연차": {"label": "연차",  "start": None, "end": None, "is_working": False},
    "공가": {"label": "공가",  "start": None, "end": None, "is_working": False},
    "병가": {"label": "병가",  "start": None, "end": None, "is_working": False},
    "OFF":  {"label": "휴무",  "start": None, "end": None, "is_working": False},
}

# ── 직종명 정규화 ─────────────────────────────────────────────────────────────
def _normalize_position(val: str) -> str:
    """줄바꿈/공백이 섞인 직종명 정규화"""
    if not val:
        return ""
    # 줄바꿈/공백 제거 후 합치기
    s = re.sub(r'[\n\r\s]+', '', str(val))
    # 알려진 직종명 매핑
    mapping = {
        "사회복지사": "사회복지사", "사회\n복지사": "사회복지사",
        "간호조무사": "간호조무사", "간호\n조무사": "간호조무사",
        "요양보호사": "요양보호사", "요\n양\n보\n호\n사": "요양보호사",
        "간호사": "간호사", "물리치료사": "물리치료사",
        "사무국장": "사무국장", "시설장": "시설장", "원장": "원장",
        "조리사": "조리사", "영양사": "영양사", "위생원": "위생원",
    }
    for raw, normalized in mapping.items():
        if s == re.sub(r'[\n\r\s]+', '', raw):
            return normalized
    return s


def _is_team_name(val: str) -> bool:
    """A조/B조/C조 등 조 이름 여부"""
    if not val:
        return False
    s = str(val).strip()
    return bool(re.match(r'^[A-Za-z가-힣]조$', s)) or s in ("A조","B조","C조","D조","E조","1조","2조","3조")


def _is_position(val: str) -> bool:
    """직종명인지 여부"""
    positions = {
        "요양보호사","간호사","간호조무사","사회복지사","물리치료사",
        "작업치료사","시설장","원장","사무국장","조리사","영양사","위생원",
        "관리자","직종",
    }
    normalized = re.sub(r'[\n\r\s]+', '', str(val or ""))
    return normalized in {re.sub(r'[\n\r\s]+', '', p) for p in positions}


def _normalize_shift(val: Any) -> Optional[Dict]:
    """셀 값을 shift 정보로 변환. 공란이면 None 반환"""
    if val is None:
        return None
    s = str(val).strip()
    if not s or s in ("", "-", "0"):
        return None
    # 대소문자 무시로 매핑
    for key, info in SHIFT_MAP.items():
        if s.upper() == key.upper() or s == key:
            return {"shift_code": key, **info}
    # 부분 매칭 (D주간 등)
    for key, info in SHIFT_MAP.items():
        if key in s or s in key:
            return {"shift_code": key, **info}
    return None


# ── 파일 읽기 ─────────────────────────────────────────────────────────────────
def _read_xlsx(file_bytes: bytes) -> List[List[Any]]:
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=False, data_only=True)
    ws = wb.active

    # 병합 셀 해제 — 병합 범위의 값을 첫 셀에서 복사
    merge_vals = {}
    for mr in list(ws.merged_cells.ranges):
        top_val = ws.cell(mr.min_row, mr.min_col).value
        for r in range(mr.min_row, mr.max_row + 1):
            for c in range(mr.min_col, mr.max_col + 1):
                merge_vals[(r, c)] = top_val

    rows = []
    for r_idx, row in enumerate(ws.iter_rows(), 1):
        row_vals = []
        for c_idx, cell in enumerate(row, 1):
            v = merge_vals.get((r_idx, c_idx), cell.value)
            row_vals.append(v)
        # 완전 빈 행 스킵
        if any(v is not None and str(v).strip() for v in row_vals):
            rows.append(row_vals)
    return rows


def _read_xls(file_bytes: bytes) -> List[List[Any]]:
    import xlrd
    book = xlrd.open_workbook(file_contents=file_bytes)
    ws   = book.sheet_by_index(0)

    # 병합 셀 정보
    merge_map = {}
    for crange in ws.merged_cells:
        rlo, rhi, clo, chi = crange
        val = ws.cell_value(rlo, clo)
        for r in range(rlo, rhi):
            for c in range(clo, chi):
                merge_map[(r, c)] = val

    rows = []
    for r in range(ws.nrows):
        row_vals = []
        for c in range(ws.ncols):
            v = merge_map.get((r, c), ws.cell_value(r, c))
            # xlrd 날짜 타입 처리
            if ws.cell_type(r, c) == 3:  # XL_CELL_DATE
                import xlrd as _xlrd
                try:
                    dt = _xlrd.xldate_as_datetime(float(v), book.datemode)
                    v = dt.strftime('%Y-%m-%d')
                except Exception:
                    pass
            row_vals.append(v if v != '' else None)
        if any(v is not None and str(v).strip() for v in row_vals):
            rows.append(row_vals)
    return rows


# ── 헤더 파싱 — 날짜 컬럼 인덱스 추출 ───────────────────────────────────────
def _find_date_columns(rows: List[List[Any]]) -> Tuple[int, Dict[int, int]]:
    """
    헤더 행에서 날짜 컬럼(1~31) 인덱스 찾기
    반환: (header_row_idx, {col_idx: day_number})
    """
    for row_idx, row in enumerate(rows[:5]):  # 상위 5행에서 헤더 탐색
        date_cols: Dict[int, int] = {}
        ignore_from: Optional[int] = None

        for col_idx, val in enumerate(row):
            if val is None:
                continue
            s = str(val).strip()

            # 무시 컬럼 키워드 만나면 이후 컬럼 무시
            if any(kw in s for kw in IGNORE_COL_KEYWORDS):
                if ignore_from is None:
                    ignore_from = col_idx
                continue

            # 1~31 숫자이면 날짜 컬럼
            try:
                day = int(float(s))
                if 1 <= day <= 31:
                    if ignore_from is None:
                        date_cols[col_idx] = day
            except (ValueError, TypeError):
                pass

        # 날짜 컬럼이 5개 이상이면 이 행이 헤더
        if len(date_cols) >= 5:
            return row_idx, date_cols

    return -1, {}


# ── 메인 파서 ─────────────────────────────────────────────────────────────────
def parse_schedule(
    file_bytes: bytes,
    filename: str,
    year: Optional[int],
    month: Optional[int],
) -> Dict:
    """
    근무표 엑셀을 파싱해서 표준 레코드 리스트 반환
    반환: {"rows": [...], "warnings": [], "openai_used": False}
    """
    ext = filename.rsplit('.', 1)[-1].lower()

    try:
        if ext == 'xlsx':
            raw_rows = _read_xlsx(file_bytes)
        elif ext == 'xls':
            raw_rows = _read_xls(file_bytes)
        else:
            return {"rows": [], "warnings": [f"근무표는 xlsx/xls만 지원합니다: {ext}"], "openai_used": False}
    except Exception as e:
        return {"rows": [], "warnings": [f"파일 읽기 실패: {e}"], "openai_used": False}

    # 헤더에서 날짜 컬럼 찾기
    header_row_idx, date_cols = _find_date_columns(raw_rows)
    if header_row_idx < 0:
        logger.warning("날짜 컬럼을 찾지 못함 — OpenAI fallback 필요")
        return {"rows": [], "warnings": ["날짜 컬럼(1~31)을 헤더에서 찾지 못했습니다."], "openai_used": False, "needs_openai": True}

    if not year or not month:
        return {"rows": [], "warnings": ["year, month 파라미터가 필요합니다."], "openai_used": False}

    records: List[Dict] = []
    warnings: List[str] = []

    current_position = ""
    current_team     = ""

    # 헤더 다음 행부터 데이터 파싱 (요일 행 1~2개 스킵)
    data_start = header_row_idx + 1
    # 요일 행 스킵 (월화수목금토일 패턴)
    weekdays = {"월","화","수","목","금","토","일"}
    for skip_idx in range(data_start, min(data_start + 3, len(raw_rows))):
        row_vals = [str(v).strip() for v in raw_rows[skip_idx] if v is not None]
        weekday_count = sum(1 for v in row_vals if v in weekdays)
        if weekday_count >= 3:
            data_start = skip_idx + 1
            break

    # 데이터 행 파싱
    for row in raw_rows[data_start:]:
        if not any(v is not None and str(v).strip() for v in row):
            continue

        # 첫 두 셀로 직종/이름 판단
        col0 = str(row[0]).strip() if row[0] is not None else ""
        col1 = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
        col2 = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ""

        col0_clean = re.sub(r'[\n\r\s]+', '', col0)

        # 직종 업데이트
        if _is_position(col0_clean) or _is_position(col0):
            current_position = _normalize_position(col0)
            current_team = ""
            # col1이 이름인지 확인
            if col1 and not _is_position(col1) and not _is_team_name(col1):
                staff_name = col1
            elif col1 and _is_team_name(col1) and col2:
                current_team = col1
                staff_name = col2
            else:
                continue  # 직종행 자체 (이름 없음)

        elif _is_team_name(col0):
            # A조/B조 등 — 다음 셀이 직원명
            current_team = col0
            if col1:
                staff_name = col1
            else:
                continue

        elif col0 and not _is_position(col0):
            # col0 자체가 직원명인 경우
            staff_name = col0

        else:
            continue

        staff_name = staff_name.strip()
        if not staff_name or len(staff_name) < 2:
            continue

        # 날짜 컬럼별 shift 파싱
        for col_idx, day in date_cols.items():
            if col_idx >= len(row):
                continue
            cell_val = row[col_idx]
            shift = _normalize_shift(cell_val)
            if shift is None:
                continue  # 공란은 저장 안 함

            try:
                work_date = f"{year}-{month:02d}-{day:02d}"
            except Exception:
                continue

            records.append({
                "staff_name":  staff_name,
                "position":    current_position,
                "team":        current_team,
                "work_date":   work_date,
                "shift_code":  shift["shift_code"],
                "shift_label": shift["label"],
                "start_time":  shift["start"],
                "end_time":    shift["end"],
                "is_working":  shift["is_working"],
                "raw_data":    {"cell": str(cell_val), "row_col0": col0, "row_col1": col1},
            })

    logger.info(f"근무표 파싱 완료: {len(records)}건 (직원×날짜)")
    return {
        "rows":        records,
        "warnings":    warnings,
        "openai_used": False,
    }
