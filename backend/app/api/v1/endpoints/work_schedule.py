"""근무표 API — 읽기·쓰기 모두 ADMIN·시설장 전용"""
from __future__ import annotations
import logging
import re
from typing import Optional, Dict, Any, List
from fastapi import APIRouter, Depends, HTTPException, Query
from pydantic import BaseModel
from sqlalchemy.orm import Session

from app.core.database import get_db
from app.core.security import get_current_user
from app.models.user import User
from app.models.work_schedule import (WorkSchedule, WorkScheduleVersion,
                                      WorkScheduleConfig, now_kst)
from app.models.eval import LtcStaffMember
from app.services.staff_notify import notify_all_staff
from app.models.staffing import HolidayCalendar
from app.schemas.response import ApiResponse
from app.services import shift_hours as _shift_hours_mod

logger = logging.getLogger(__name__)
router = APIRouter()
_YM = re.compile(r"^\d{4}-\d{2}$")


def _manager(current_user: User = Depends(get_current_user)) -> User:
    role = current_user.role.value if hasattr(current_user.role, "value") else str(current_user.role)
    pos = getattr(current_user, "position", None)
    pos = pos.value if hasattr(pos, "value") else str(pos or "")
    if role != "ADMIN" and pos != "시설장":
        raise HTTPException(403, "근무표 접근 권한이 없습니다. (관리자·시설장)")
    return current_user


def _viewer(current_user: User = Depends(get_current_user)) -> User:
    """읽기 전용 — 대표·이사도 전체 근무표를 볼 수 있다(수정은 _manager만)."""
    role = current_user.role.value if hasattr(current_user.role, "value") else str(current_user.role)
    pos = getattr(current_user, "position", None)
    pos = pos.value if hasattr(pos, "value") else str(pos or "")
    if role != "ADMIN" and pos not in ("시설장", "대표", "이사"):
        raise HTTPException(403, "근무표 열람 권한이 없습니다.")
    return current_user


def _prev_ym(ym: str) -> str:
    y, m = int(ym[:4]), int(ym[5:7])
    m -= 1
    if m == 0:
        m, y = 12, y - 1
    return f"{y}-{m:02d}"


def _inherit_rows(db: Session, ym: str, back: int = 6) -> tuple:
    """이번 달에 조 편성이 없으면 가장 가까운 이전 달 것을 물려준다.
    매달 조를 다시 짜는 건 낭비이고, 조 구성은 잘 바뀌지 않기 때문."""
    cur = ym
    for _ in range(back):
        cur = _prev_ym(cur)
        w = db.query(WorkSchedule).filter(WorkSchedule.year_month == cur).first()
        if w and w.rows:
            return w.rows, cur
    return [], None


def _view(w: Optional[WorkSchedule], ym: str, db: Optional[Session] = None) -> dict:
    return {
        "year_month": ym,
        "data": (w.data if w and w.data else {}),
        "rows": (w.rows if w and w.rows else []),
        # 저장된 값이 없으면 null — 프론트가 (토·일·공휴일 제외 일수)×8로 자동 계산한다.
        # 예전에는 여기서 "160"을 내려보내 자동 계산값을 덮어썼다.
        "base_hours": (w.base_hours if w else None) or None,
        "base_days": (w.base_days if w else None) or None,
        "as_of": (w.as_of if w else None),
        "team_offsets": (w.team_offsets if w and w.team_offsets else None),
        "updated_by": (w.updated_by if w else None),
        "updated_at": (w.updated_at.isoformat() if w and w.updated_at else None),
        "locked": bool(w and w.locked_at),
        "locked_by": (w.locked_by if w else None),
        "locked_at": (w.locked_at.isoformat() if w and w.locked_at else None),
    }


KEEP_VERSIONS = 30      # 월별로 최근 30개까지만 보관


def _count_cells(data: Optional[Dict[str, Any]]) -> int:
    """입력된 근무 칸 수"""
    return sum(1 for row in (data or {}).values() if isinstance(row, dict)
               for v in row.values() if v)


def _diff_cells(a: Optional[Dict[str, Any]], b: Optional[Dict[str, Any]]) -> int:
    """두 근무표 사이에 값이 달라진 칸 수 (추가·삭제·변경 모두)"""
    a, b = a or {}, b or {}
    keys = set()
    for sid in set(a) | set(b):
        for day in set((a.get(sid) or {})) | set((b.get(sid) or {})):
            keys.add((sid, day))
    return sum(1 for sid, day in keys
               if (a.get(sid) or {}).get(day) != (b.get(sid) or {}).get(day))


def _version_view(v: WorkScheduleVersion, full: bool = False) -> dict:
    out = {
        "id": v.id, "year_month": v.year_month,
        "cells": v.cells or 0, "changed": v.changed or 0,
        "base_hours": v.base_hours, "base_days": v.base_days,
        "saved_by": v.saved_by, "saved_at": v.saved_at.isoformat() if v.saved_at else None,
    }
    if full:
        out.update({"data": v.data or {}, "rows": v.rows or [],
                    "as_of": v.as_of, "team_offsets": v.team_offsets})
    return out


class ScheduleBody(BaseModel):
    year_month: str
    data: Dict[str, Any] = {}
    rows: Optional[List[Dict[str, Any]]] = None
    base_hours: Optional[str] = None
    base_days: Optional[str] = None
    as_of: Optional[str] = None
    team_offsets: Optional[Dict[str, int]] = None


class ConfigBody(BaseModel):
    settle_start: Optional[str] = None
    rotation_anchor: Optional[str] = None
    # {'N': 10, …} — 비운 코드는 기본값으로 되돌아간다
    code_hours: Optional[Dict[str, float]] = None


def _config_row(db: Session) -> WorkScheduleConfig:
    row = db.query(WorkScheduleConfig).first()
    if not row:
        row = WorkScheduleConfig(settle_start="2026-07", rotation_anchor="2026-08-01")
        db.add(row); db.commit(); db.refresh(row)
    return row


@router.post("/notify")
def notify_schedule(body: ScheduleBody, db: Session = Depends(get_db),
                    current_user: User = Depends(_manager)):
    """근무표 발표 알림 — 전 직원 푸시. 누르면 직원앱이 '내 근무표'를 연다."""
    if not _YM.match(body.year_month or ""):
        raise HTTPException(400, "year_month 형식은 YYYY-MM 이어야 합니다.")
    y, m = body.year_month.split("-")
    result = notify_all_staff(
        db,
        f"{int(m)}월 근무표가 나왔습니다",
        "내 근무표에서 이번 달 근무를 확인하세요.",
        data={"type": "my-schedule", "month": body.year_month},
        exclude_user_id=getattr(current_user, "id", None),
    )
    return ApiResponse(success=True, data=result)


@router.get("/mine")
def my_schedule(month: str = Query(...), db: Session = Depends(get_db),
                current_user: User = Depends(get_current_user)):
    """로그인한 직원 본인의 한 달 근무 — 관리자 권한 없이 전 직원이 본다.

    계정 관리에서 연동한 직원을 최우선으로, 미연동 계정은 이름 매칭으로 찾는다."""
    if not _YM.match(month or ""):
        raise HTTPException(400, "month 형식은 YYYY-MM 이어야 합니다.")
    from app.services.staff_link import resolve_staff_for_user
    staff = resolve_staff_for_user(db, current_user)

    w = db.query(WorkSchedule).filter(WorkSchedule.year_month == month).first()
    codes = (w.data or {}).get(staff.id, {}) if w else {}
    team = None
    for r in (w.rows or []) if w else []:
        if r.get("staff_id") == staff.id:
            team = r.get("team")
            break
    return ApiResponse(success=True, data={
        "year_month": month, "staff_name": staff.name, "team": team,
        "codes": codes,
        # 개인별 한 줄 설명 — 저장 시 생성. "왜 이렇게 나왔어요?"에 대한 답
        "note": ((w.notes or {}).get(staff.id) if w else None),
        "updated_at": (w.updated_at.isoformat() if w and w.updated_at else None),
    })


@router.get("/config")
def get_config(db: Session = Depends(get_db), _: User = Depends(_manager)):
    """정산 시작월·회전 기준일 — 연도가 바뀌면 여기만 고치면 된다."""
    row = _config_row(db)
    return ApiResponse(success=True, data={
        "settle_start": row.settle_start or "2026-07",
        "rotation_anchor": row.rotation_anchor or "2026-08-01",
        # 코드별 시간 — 비어 있으면 화면이 기본값을 쓴다
        "code_hours": row.code_hours or {},
        # 무엇을 고칠 수 있는지 화면이 알아야 목록을 그린다
        "code_hours_default": dict(_shift_hours_mod.CODE_HOURS),
    })


@router.put("/config")
def save_config(body: ConfigBody, db: Session = Depends(get_db), current_user: User = Depends(_manager)):
    if body.settle_start and not re.match(r"^\d{4}-\d{2}$", body.settle_start):
        raise HTTPException(400, "정산 시작월은 YYYY-MM 형식이어야 합니다.")
    if body.rotation_anchor and not re.match(r"^\d{4}-\d{2}-\d{2}$", body.rotation_anchor):
        raise HTTPException(400, "회전 기준일은 YYYY-MM-DD 형식이어야 합니다.")
    row = _config_row(db)
    if body.settle_start is not None: row.settle_start = body.settle_start or None
    if body.rotation_anchor is not None: row.rotation_anchor = body.rotation_anchor or None
    if body.code_hours is not None:
        # 아는 코드만, 말이 되는 범위만 받는다. 근무표는 급여로 이어지므로
        # 오타 하나가 한 달치 계산을 바꾼다.
        clean = {}
        for k, v in body.code_hours.items():
            if k not in _shift_hours_mod.CODE_HOURS:
                raise HTTPException(400, f"알 수 없는 근무 코드입니다: {k}")
            try:
                fv = float(v)
            except (TypeError, ValueError):
                raise HTTPException(400, f"{k} 시간이 숫자가 아닙니다.")
            if not (0 <= fv <= 24):
                raise HTTPException(400, f"{k} 시간은 0~24 사이여야 합니다.")
            # 기본값과 같으면 담아둘 이유가 없다 — 나중에 기본값을 고치면
            # 굳어 있던 값이 발목을 잡는다
            if abs(fv - _shift_hours_mod.CODE_HOURS[k]) > 1e-9:
                clean[k] = fv
        row.code_hours = clean or None
    row.updated_by = getattr(current_user, "name", None)
    db.commit(); db.refresh(row)
    return ApiResponse(success=True, data={
        "code_hours": row.code_hours or {},
        "settle_start": row.settle_start, "rotation_anchor": row.rotation_anchor,
    })


@router.get("/export")
def export_schedule(month: str = Query(...), db: Session = Depends(get_db), _: User = Depends(_viewer)):
    """저장된 근무표 최종본을 엑셀(.xlsx)로 — 기존 수기 근무표 서식을 따른다.

    참고 원본: 노란 제목띠 · 직종/조/성명 열 · 주말 색(토 파랑/일 빨강) ·
    休 파랑 배경 · 대휴 주황 배경 · 요양보호사 직종 세로 병합 ·
    하단 일별 근무 인원 합계 + 근무 코드 범례.
    """
    if not _YM.match(month or ""):
        raise HTTPException(400, "month 형식은 YYYY-MM 이어야 합니다.")
    w = db.query(WorkSchedule).filter(WorkSchedule.year_month == month).first()
    if not w or not (w.data or {}):
        raise HTTPException(404, "저장된 근무표가 없습니다.")

    import io as _io
    import calendar as _cal
    from datetime import date as _date
    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    from app.services import shift_hours as _shift_hours

    # 코드별 시간 설정 — 화면과 같은 값으로 계산해야 한다.
    # 설정을 안 읽으면 야간을 10시간으로 바꿔둬도 엑셀만 9시간으로 센다.
    _code_hours = (_config_row(db).code_hours or {})

    y, m = int(month[:4]), int(month[5:7])
    total = _cal.monthrange(y, m)[1]
    names = {s2.id: s2.name for s2 in db.query(LtcStaffMember).all()}
    poss = {s2.id: (s2.position or "") for s2 in db.query(LtcStaffMember).all()}

    # 공휴일 (빨간 날만)
    holidays = set()
    try:
        from app.services import staffing as S
        rows_h = []
        try:
            rows_h = [{"date": r.date, "name": r.name, "kind": r.kind}
                      for r in db.query(HolidayCalendar).filter(HolidayCalendar.active == True).all()]  # noqa: E712
        except Exception:
            pass
        kinds = {r["date"]: (r.get("kind") or "public") for r in rows_h}
        for d, n in (S.get_korean_holidays(y, None, rows_h) or {}).items():
            if d.startswith(month) and not (n in ("근로자의 날",) or kinds.get(d) == "paid"):
                holidays.add(int(d[8:10]))
    except Exception:
        pass

    wb = Workbook()
    ws = wb.active
    ws.title = f"{y % 100}.{m:02d}월"

    # ── 원본 서식 상수 ──
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    F = lambda **kw: Font(name="Arial", size=12, bold=True, **kw)   # noqa: E731
    RED, BLUE = "FF0000", "0000FF"
    SAT_FILL = PatternFill("solid", fgColor="DAEEF3")    # 토 옅은 파랑
    SUN_FILL = PatternFill("solid", fgColor="F4CCCC")    # 일·공휴일 옅은 빨강
    HYU_FILL = PatternFill("solid", fgColor="4285F4")    # 休 파랑
    DH_FILL = PatternFill("solid", fgColor="FF6D01")     # 대휴 주황
    TITLE_FILL = PatternFill("solid", fgColor="FFFF00")  # 제목 노랑

    # C열 성명 · D열 총시간 · E열부터 날짜
    #
    # 총시간은 저장할 때 화면이 계산해 담아 보낸 값을 그대로 쓴다.
    # 여기서 다시 계산하지 않는다 — 근무시간 계산에는 휴게시간 규칙과
    # 직접 입력한 시간대 처리가 얽혀 있어, 파이썬에 옮겨 적으면 언젠가
    # 화면과 엑셀의 숫자가 갈라진다. 그러면 어느 쪽이 맞는지 아무도 모른다.
    NAME_C, TOTAL_C, DAY0 = 3, 4, 5

    def day_style(c, day, base_fill=None):
        dow = _date(y, m, day).weekday()
        red = dow == 6 or day in holidays
        if red:
            c.fill = SUN_FILL
        elif dow == 5:
            c.fill = SAT_FILL
        elif base_fill:
            c.fill = base_fill
        return red, dow == 5

    # ── 1행: 제목 (노란 띠) ──
    ws.merge_cells(start_row=1, start_column=DAY0, end_row=1, end_column=DAY0 + total - 1)
    tc = ws.cell(row=1, column=DAY0, value=f"{y}년  {m}월 행복한요양원 근무표")
    tc.font = Font(name="맑은 고딕", size=14, bold=True)
    tc.fill = TITLE_FILL
    tc.alignment = center
    tc.border = border
    ws.row_dimensions[1].height = 25.5

    # ── 3·4행: 날짜·요일 머리 ──
    ws.merge_cells("A3:B4")
    ws.merge_cells("C3:C4")
    h1 = ws.cell(row=3, column=1, value="직종"); h1.font = F(); h1.alignment = center; h1.border = border
    ws.cell(row=4, column=1).border = border
    ws.cell(row=3, column=2).border = border
    ws.cell(row=4, column=2).border = border
    h2 = ws.cell(row=3, column=NAME_C, value="성명"); h2.font = F(); h2.alignment = center; h2.border = border
    ws.cell(row=4, column=NAME_C).border = border

    # 총시간 — 기준 근로시간을 아래 칸에 함께 적어 견줄 수 있게
    base_h = (w.base_hours or "").strip()
    ws.cell(row=3, column=TOTAL_C, value="총시간").font = F()
    ws.cell(row=3, column=TOTAL_C).alignment = center
    ws.cell(row=3, column=TOTAL_C).border = border
    bh = ws.cell(row=4, column=TOTAL_C, value=(f"기준 {base_h}" if base_h else ""))
    bh.font = Font(name="Arial", size=9, bold=False)
    bh.alignment = center
    bh.border = border
    DOW_KO = ["월", "화", "수", "목", "금", "토", "일"]
    for day in range(1, total + 1):
        col = DAY0 + day - 1
        c1 = ws.cell(row=3, column=col, value=day)
        c2 = ws.cell(row=4, column=col, value=DOW_KO[_date(y, m, day).weekday()])
        for c in (c1, c2):
            c.alignment = center
            c.border = border
            red, sat = day_style(c, day)
            c.font = F(color=RED) if red else F(color=BLUE) if sat else F()
    ws.row_dimensions[4].height = 30

    # ── 본문 — 저장된 rows 순서(화면 정렬 그대로) ──
    body_rows = []
    for row in (w.rows or []):
        sid = row.get("staff_id")
        codes = (w.data or {}).get(sid) or {}
        if not codes and sid not in names:
            continue
        body_rows.append((row, sid, codes))

    r_i = 5
    cg_rows = []          # 요양보호사 블록(직종 세로 병합용)
    care_first = None
    for row, sid, codes in body_rows:
        pos = row.get("position") or poss.get(sid, "")
        is_care = "요양보호사" in pos
        # 직종 (요양보호사는 나중에 병합, 그 외는 A:B 가로 병합)
        if is_care:
            if care_first is None:
                care_first = r_i
            cg_rows.append(r_i)
            tv = ws.cell(row=r_i, column=2, value=row.get("team") or "")
            tv.font = F(); tv.alignment = center
        else:
            ws.merge_cells(start_row=r_i, start_column=1, end_row=r_i, end_column=2)
            pv = ws.cell(row=r_i, column=1, value=pos)
            pv.font = F(); pv.alignment = center
        nv = ws.cell(row=r_i, column=NAME_C, value=names.get(sid, "(퇴사)"))
        nv.font = F(); nv.alignment = center

        # 총시간 — 저장할 때 화면이 계산해 담아 보낸 값을 먼저 쓴다.
        #
        # 이 기능이 생기기 전에 저장된 달에는 그 값이 없다. 급여로 이어지는
        # 숫자를 '한 번 저장하셔야 나옵니다' 로 둘 수는 없어, 없으면 여기서
        # 계산한다. 계산 규칙은 화면과 같은 표를 통과하는 것을 테스트가
        # 확인한다(backend/tests/test_shift_hours.py).
        tv_raw = row.get("total")
        if not isinstance(tv_raw, (int, float)):
            tv_raw = _shift_hours.month_total(codes, range(1, total + 1), _code_hours)
        tc2 = ws.cell(row=r_i, column=TOTAL_C,
                      value=(float(tv_raw) if isinstance(tv_raw, (int, float)) else None))
        tc2.font = F()
        tc2.alignment = center
        tc2.number_format = "0.#"
        try:
            _b = float(base_h) if base_h else 0.0
        except ValueError:
            _b = 0.0
        if isinstance(tv_raw, (int, float)) and _b > 0:
            # 미달은 급여가 깎이는 쪽이라 더 급하다. 초과는 수당 문제다.
            if float(tv_raw) < _b:
                tc2.font = F(color=RED)
            elif float(tv_raw) > _b:
                tc2.font = F(color="E36C09")
        for cix in (1, 2, NAME_C, TOTAL_C):
            ws.cell(row=r_i, column=cix).border = border
        for day in range(1, total + 1):
            code = codes.get(str(day), "")
            # "0930 1200" 같은 시간대 근무는 두 줄로 (원본과 동일)
            disp = code
            import re as _re2
            mm2 = _re2.match(r"^(\d{3,4})[\s~\-]+(\d{3,4})$", str(code).strip())
            if mm2:
                disp = f"{mm2.group(1)}\n{mm2.group(2)}"
            c = ws.cell(row=r_i, column=DAY0 + day - 1, value=disp)
            c.alignment = center
            c.border = border
            if code == "休":
                c.fill = HYU_FILL
                c.font = F()
            elif code in ("대휴", "초과휴"):
                c.fill = DH_FILL
                c.font = F()
            else:
                red, sat = day_style(c, day)
                c.font = F(color=RED) if (red and not code) else F()
        ws.row_dimensions[r_i].height = 34
        r_i += 1

    # 요양보호사 직종 세로 병합
    if cg_rows:
        ws.merge_cells(start_row=cg_rows[0], start_column=1, end_row=cg_rows[-1], end_column=1)
        cc = ws.cell(row=cg_rows[0], column=1, value="요\n양\n보\n호\n사")
        cc.font = F(); cc.alignment = center
        for rr in cg_rows:
            ws.cell(row=rr, column=1).border = border

    # ── 일별 근무 인원 합계 (요양보호사 기준, 원본의 특이사항 행) ──
    sum_r = r_i
    ws.merge_cells(start_row=sum_r, start_column=1, end_row=sum_r, end_column=TOTAL_C)
    sv = ws.cell(row=sum_r, column=1, value="일별 근무 인원")
    sv.font = F(); sv.alignment = center; sv.border = border
    for cix in (2, NAME_C, TOTAL_C):
        ws.cell(row=sum_r, column=cix).border = border
    WORK_CODES = {"D", "M", "N", "AD", "PD", "D-3"}
    for day in range(1, total + 1):
        n_work = 0
        for row, sid, codes in body_rows:
            code = str(codes.get(str(day), "")).strip()
            if code in WORK_CODES or (code and code[0].isdigit()):
                n_work += 1
        c = ws.cell(row=sum_r, column=DAY0 + day - 1, value=n_work)
        c.alignment = center; c.border = border
        red, sat = day_style(c, day)
        c.font = F(color=RED) if red else F(color=BLUE) if sat else F()
    r_i += 2

    # ── 범례 (원본 하단) ──
    LEGEND = [
        ("D", "08:50 ~ 18:00"), ("M", "06:50 ~ 16:00"), ("N", "17:50 ~ 익일 09:00"),
        ("AD", "08:50 ~ 13:30"), ("PD", "13:30 ~ 18:00"), ("D-3", "08:50 ~ 15:00"),
        ("休", "연차"), ("대휴", "대체휴무"), ("숫자", "시간대 근무 (직접 입력)"),
    ]
    for code, desc in LEGEND:
        cv = ws.cell(row=r_i, column=DAY0, value=code)
        cv.font = F(); cv.alignment = center; cv.border = border
        if code == "休":
            cv.fill = HYU_FILL
        elif code == "대휴":
            cv.fill = DH_FILL
        ws.merge_cells(start_row=r_i, start_column=DAY0 + 1, end_row=r_i, end_column=DAY0 + 5)
        dv = ws.cell(row=r_i, column=DAY0 + 1, value=desc)
        dv.font = Font(name="Arial", size=11)
        dv.alignment = Alignment(horizontal="left", vertical="center")
        for cix in range(DAY0 + 1, DAY0 + 6):
            ws.cell(row=r_i, column=cix).border = border
        r_i += 1

    # ── 폭·틀 고정 ──
    ws.column_dimensions["A"].width = 5.5
    ws.column_dimensions["B"].width = 6.4
    ws.column_dimensions["C"].width = 8.4
    ws.column_dimensions["D"].width = 7.2      # 총시간
    for d in range(1, total + 1):
        ws.column_dimensions[get_column_letter(DAY0 + d - 1)].width = 6.5
    ws.freeze_panes = "D5"

    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    # 파일명: 원본 시트명 관례를 따라 "26.08월 (26.08.08).xlsx" — 괄호 안은 만든 날짜
    from datetime import datetime as _dtm, timezone as _tz, timedelta as _td2
    today_kst = _dtm.now(_tz(_td2(hours=9)))
    fname = f"{y % 100:02d}.{m:02d}월 ({today_kst.strftime('%y.%m.%d')}).xlsx"
    from urllib.parse import quote as _q
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{_q(fname)}"})


@router.get("/holidays")
def month_holidays(month: str = Query(...), db: Session = Depends(get_db), _: User = Depends(_viewer)):
    """해당 월의 공휴일 { 'YYYY-MM-DD': {name, kind} }.

    kind='paid' = 근로자의 날처럼 관공서 공휴일은 아니지만 유급휴일인 날.
    빨간 날과 색·계산을 다르게 다루기 위해 종류를 함께 내려준다."""
    if not _YM.match(month or ""):
        raise HTTPException(400, "month 형식은 YYYY-MM 이어야 합니다.")
    y = int(month[:4])

    # DB에 등록된 공휴일 (음력 명절·대체공휴일·임시공휴일)
    table: List[Dict[str, Any]] = []
    try:
        rows = db.query(HolidayCalendar).filter(HolidayCalendar.active == True).all()  # noqa: E712
        table = [{"date": r.date, "name": r.name, "kind": r.kind} for r in rows]
    except Exception as e:
        logger.warning("holiday_calendar 조회 실패: %s", e)

    # 라이브러리/규칙 기반 공휴일과 병합 (실패해도 DB 값은 살린다)
    try:
        from app.services import staffing as S
        hol = S.get_korean_holidays(y, None, table)
    except Exception as e:
        logger.warning("공휴일 계산 실패, DB 값만 사용: %s", e)
        hol = {r["date"]: (r.get("name") or "공휴일") for r in table}

    kinds = {r["date"]: (r.get("kind") or "public") for r in table}
    PAID = {"근로자의 날"}
    out = {}
    for d, n in (hol or {}).items():
        if not d.startswith(month):
            continue
        k = kinds.get(d) or "public"
        if n in PAID or k == "paid":
            k = "paid"
        out[d] = {"name": n, "kind": k}
    return ApiResponse(success=True, data=out)


@router.get("")
def get_schedule(month: str = Query(...), db: Session = Depends(get_db), _: User = Depends(_viewer)):
    if not _YM.match(month or ""):
        raise HTTPException(400, "month 형식은 YYYY-MM 이어야 합니다.")
    w = db.query(WorkSchedule).filter(WorkSchedule.year_month == month).first()
    data = _view(w, month)
    # 이번 달 조 편성이 비어 있으면 이전 달 것을 이어받는다
    if not data["rows"]:
        rows, src = _inherit_rows(db, month)
        if rows:
            data["rows"] = rows
            data["rows_from"] = src        # 어느 달에서 가져왔는지 화면에 알려준다
    return ApiResponse(success=True, data=data)


class ExplainPerson(BaseModel):
    staff_id: str
    name: str
    team: Optional[str] = None
    hours: int = 0            # 총시간 (추가근무 포함)
    base: int = 0             # 월 기준시간
    d: int = 0                # 주간 근무 수
    n: int = 0                # 야간 근무 수
    annual: int = 0           # 연차(休)
    daehyu: int = 0           # 대휴
    comp: int = 0             # 초과휴
    extra: int = 0            # 추가근무 시간
    carry: Optional[int] = None   # 다음 달로 이월되는 미상환 시간


class ExplainBody(BaseModel):
    month: str
    people: List[ExplainPerson]


def _fallback_note(p: ExplainPerson) -> str:
    """AI 없이도 나오는 기본 한 줄 — 숫자는 어차피 여기 다 있다."""
    bits = []
    if p.n: bits.append(f"주간 {p.d}·야간 {p.n}")
    elif p.d: bits.append(f"주간 {p.d}일")
    if p.annual: bits.append(f"연차 {p.annual}일")
    if p.daehyu: bits.append(f"공휴일 근무 보상 대휴 {p.daehyu}일")
    if p.comp: bits.append(f"밀린 추가근무 보상 초과휴 {p.comp}일")
    if p.extra: bits.append(f"추가근무 {p.extra}시간")
    body = ", ".join(bits) if bits else "이번 달 근무"
    return f"{body} — 총 {p.hours}시간 (기준 {p.base}시간 충족)"


@router.post("/explain")
def explain_schedule(body: ExplainBody, db: Session = Depends(get_db),
                     current_user: User = Depends(_manager)):
    """저장된 근무표에 개인별 한 줄 설명을 붙인다.

    "왜 나는 이번 달 이렇게 나왔어요?"에 관리자가 일일이 답하지 않도록,
    정산 숫자를 사람 말로 풀어 직원 내 근무표에 보여준다.
    계산은 프론트 정산 엔진 결과를 그대로 받고, AI는 문장만 만든다(실패 시 템플릿)."""
    w = db.query(WorkSchedule).filter(WorkSchedule.year_month == body.month).first()
    if not w:
        raise HTTPException(404, "먼저 근무표를 저장해주세요.")

    notes = {p.staff_id: _fallback_note(p) for p in body.people}
    ai_used = False
    try:
        from app.core.config import settings
        if settings.OPENAI_API_KEY and body.people:
            import json
            from openai import OpenAI
            client = OpenAI(api_key=settings.OPENAI_API_KEY)
            lines = "\n".join(
                f"- id={p.name}: 총 {p.hours}h/기준 {p.base}h, 주간 {p.d}, 야간 {p.n}, "
                f"연차 {p.annual}, 대휴 {p.daehyu}, 초과휴 {p.comp}, 추가근무 {p.extra}h"
                + (f", 이월 {p.carry}h" if p.carry else "")
                for p in body.people)
            r = client.chat.completions.create(
                model=settings.OPENAI_MODEL or "gpt-4o-mini",
                messages=[{"role": "user", "content": (
                    "요양원 근무표를 받은 50~60대 요양보호사 선생님께 보여줄 개인별 한 줄 설명을 만들어주세요.\n"
                    "규칙: 각자 60자 이내 존댓말 1문장. 숫자를 바꾸거나 새로 만들지 말 것. "
                    "대휴=공휴일 근무 보상, 초과휴=밀린 추가근무 보상이라는 취지가 드러나게. "
                    "쉬운 말로, 따뜻하지만 담백하게.\n"
                    f"{body.month} 근무 내역:\n{lines}\n\n"
                    '출력은 JSON 하나만: {"이름": "설명", ...}'
                )}],
                max_tokens=1500, temperature=0.4, timeout=30,
            )
            txt = (r.choices[0].message.content or "").strip()
            txt = txt[txt.find("{"): txt.rfind("}") + 1]
            by_name = json.loads(txt)
            name_to_id = {p.name: p.staff_id for p in body.people}
            for nm, note in by_name.items():
                sid = name_to_id.get(nm)
                if sid and isinstance(note, str) and 5 <= len(note) <= 120:
                    notes[sid] = note.strip()
            ai_used = True
    except Exception as e:
        logger.warning("근무표 설명 AI 생성 실패 — 템플릿 사용: %s", e)

    w.notes = notes                      # JSON 재할당으로 변경 감지
    db.commit()
    return ApiResponse(success=True, data={"count": len(notes), "ai": ai_used})


@router.put("")
def save_schedule(body: ScheduleBody, db: Session = Depends(get_db), current_user: User = Depends(_manager)):
    if not _YM.match(body.year_month or ""):
        raise HTTPException(400, "year_month 형식은 YYYY-MM 이어야 합니다.")
    guard_locked(db, body.year_month)
    w = db.query(WorkSchedule).filter(WorkSchedule.year_month == body.year_month).first()
    if not w:
        w = WorkSchedule(year_month=body.year_month)
        db.add(w)
    w.data = body.data or {}
    if body.rows is not None: w.rows = body.rows
    if body.base_hours is not None: w.base_hours = body.base_hours
    if body.base_days is not None: w.base_days = body.base_days
    if body.as_of is not None: w.as_of = body.as_of
    if body.team_offsets is not None: w.team_offsets = body.team_offsets
    w.updated_by = getattr(current_user, "name", None)

    # 저장 시점 스냅샷을 남긴다 — 편성이 꼬였을 때 되돌리기 위한 이력
    try:
        prev = (db.query(WorkScheduleVersion)
                .filter(WorkScheduleVersion.year_month == body.year_month)
                .order_by(WorkScheduleVersion.saved_at.desc()).first())
        changed = _diff_cells(prev.data if prev else {}, w.data)
        # 근무 칸이 하나도 안 바뀌었으면 이력을 새로 만들지 않는다 (기준값만 고친 경우 등)
        if prev is None or changed > 0:
            db.add(WorkScheduleVersion(
                year_month=body.year_month, data=w.data, rows=w.rows,
                base_hours=w.base_hours, base_days=w.base_days,
                as_of=w.as_of, team_offsets=w.team_offsets,
                cells=_count_cells(w.data), changed=changed,
                saved_by=getattr(current_user, "name", None),
            ))
            db.flush()
            # 오래된 이력 정리
            olds = (db.query(WorkScheduleVersion)
                    .filter(WorkScheduleVersion.year_month == body.year_month)
                    .order_by(WorkScheduleVersion.saved_at.desc())
                    .offset(KEEP_VERSIONS).all())
            for o in olds:
                db.delete(o)
    except Exception as e:
        logger.warning("근무표 이력 기록 실패(저장은 계속): %s", e)

    db.commit(); db.refresh(w)
    return ApiResponse(success=True, data=_view(w, body.year_month))


# ── 확정 잠금 ──────────────────────────────────────────────────────────────
# 근무표는 붙여 놓고 여러 사람이 보는 문서다. 확정한 뒤 조용히 바뀌면
# 사람마다 다른 표를 보게 되고, 그날 누가 나오는지가 어긋난다.
# 그래서 잠글 수 있게 하고, 잠근 달은 어디에서도 못 고치게 막는다.

def schedule_locked(db: Session, ym: str) -> Optional[str]:
    """잠겨 있으면 사람이 읽을 안내 문구, 아니면 None."""
    w = db.query(WorkSchedule).filter(WorkSchedule.year_month == ym).first()
    if not w or not w.locked_at:
        return None
    who = f" ({w.locked_by})" if w.locked_by else ""
    return (f"{int(ym[5:7])}월 근무표는 확정되어 잠겨 있습니다{who}. "
            f"고치려면 근무표 화면에서 잠금을 풀어주세요. (ADMIN만 가능)")


def guard_locked(db: Session, ym: str) -> None:
    msg = schedule_locked(db, ym)
    if msg:
        raise HTTPException(409, msg)


class LockBody(BaseModel):
    year_month: str
    locked: bool


@router.post("/lock")
def set_lock(body: LockBody, db: Session = Depends(get_db),
             current_user: User = Depends(get_current_user)):
    """근무표 확정 잠금 — ADMIN만.

    시설장도 근무표를 짜지만, 확정 여부는 한 사람이 정해야 한다.
    """
    role = current_user.role.value if hasattr(current_user.role, "value") else str(current_user.role)
    if role != "ADMIN":
        raise HTTPException(403, "근무표 잠금은 ADMIN만 할 수 있습니다.")
    if not _YM.match(body.year_month or ""):
        raise HTTPException(400, "year_month 형식은 YYYY-MM 이어야 합니다.")
    w = db.query(WorkSchedule).filter(WorkSchedule.year_month == body.year_month).first()
    if not w:
        raise HTTPException(404, "저장된 근무표가 없습니다. 먼저 저장해주세요.")
    if body.locked:
        w.locked_at = now_kst()
        w.locked_by = getattr(current_user, "name", None)
    else:
        w.locked_at = None
        w.locked_by = None
    db.commit(); db.refresh(w)
    return ApiResponse(success=True, data=_view(w, body.year_month),
                       message="확정 잠금했습니다." if body.locked else "잠금을 풀었습니다.")


@router.get("/versions")
def list_versions(month: str = Query(...), db: Session = Depends(get_db), _: User = Depends(_manager)):
    """해당 월의 저장 이력 (최신순)"""
    if not _YM.match(month or ""):
        raise HTTPException(400, "month 형식은 YYYY-MM 이어야 합니다.")
    rows = (db.query(WorkScheduleVersion)
            .filter(WorkScheduleVersion.year_month == month)
            .order_by(WorkScheduleVersion.saved_at.desc()).limit(KEEP_VERSIONS).all())
    return ApiResponse(success=True, data=[_version_view(v) for v in rows])


@router.get("/versions/{vid}")
def get_version(vid: str, db: Session = Depends(get_db), _: User = Depends(_manager)):
    """저장 이력 하나를 통째로 — 화면에서 불러오기용"""
    v = db.query(WorkScheduleVersion).filter(WorkScheduleVersion.id == vid).first()
    if not v:
        raise HTTPException(404, "저장 이력을 찾을 수 없습니다.")
    return ApiResponse(success=True, data=_version_view(v, full=True))
