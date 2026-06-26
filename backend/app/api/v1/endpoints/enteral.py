"""
경관식(튜브 영양식) 재고 관리.
- 제품(종류) CRUD
- 입출고(입고/출고·반출) 내역 + 제품별 재고 집계
권한: ADMIN 또는 사회복지사·간호조무사·이사·대표·시설장
"""
from __future__ import annotations

import logging
from collections import defaultdict
from typing import Optional

import io
from datetime import datetime

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy.orm import Session

from app.core.database import get_db
from app.core.security import get_current_user
from app.models.user import User
from app.models.enteral import EnteralProduct, EnteralTransaction, now_kst
from app.models.carefor import CareforResident
from app.schemas.response import ApiResponse

logger = logging.getLogger("enteral")

router = APIRouter()

ALLOWED_POSITIONS = {"사회복지사", "간호조무사", "이사", "대표", "시설장"}
TX_TYPES = {"in", "out"}


def _require_manager(current_user: User = Depends(get_current_user)) -> User:
    role = current_user.role.value if hasattr(current_user.role, "value") else str(current_user.role)
    position = current_user.position.value if hasattr(current_user.position, "value") else str(current_user.position or "")
    if role != "ADMIN" and position not in ALLOWED_POSITIONS:
        raise HTTPException(status_code=403, detail="권한이 없습니다.")
    return current_user


def _product_view(p: EnteralProduct, stock: int = 0) -> dict:
    return {
        "id": p.id, "name": p.name, "brand": p.brand, "unit": p.unit,
        "spec": p.spec, "memo": p.memo, "is_active": p.is_active,
        "unit_price": p.unit_price,
        "sort_order": p.sort_order, "stock": stock,
        "created_at": p.created_at.isoformat() if p.created_at else None,
    }


def _tx_view(t: EnteralTransaction) -> dict:
    return {
        "id": t.id, "product_id": t.product_id, "product_name": t.product_name,
        "tx_type": t.tx_type, "quantity": t.quantity,
        "unit_price": t.unit_price, "amount": (t.unit_price or 0) * (t.quantity or 0),
        "resident_name": t.resident_name, "resident_id": t.resident_id,
        "tx_date": t.tx_date, "note": t.note, "created_by": t.created_by,
        "created_at": t.created_at.isoformat() if t.created_at else None,
    }


def _stock_map(db: Session) -> dict:
    """제품별 현재 재고 = 입고합 - 출고합"""
    stock = defaultdict(int)
    for t in db.query(EnteralTransaction).all():
        if not t.product_id:
            continue
        stock[t.product_id] += (t.quantity if t.tx_type == "in" else -t.quantity)
    return stock


# --------------------------------------------------------------------------- #
# 제품(종류)
# --------------------------------------------------------------------------- #
@router.get("/products")
def list_products(db: Session = Depends(get_db), current_user: User = Depends(_require_manager)):
    stock = _stock_map(db)
    rows = db.query(EnteralProduct).order_by(EnteralProduct.sort_order.asc(), EnteralProduct.name.asc()).all()
    return ApiResponse(success=True, data=[_product_view(p, stock.get(p.id, 0)) for p in rows])


class ProductBody(BaseModel):
    name: str
    brand: Optional[str] = None
    unit: Optional[str] = None
    spec: Optional[str] = None
    memo: Optional[str] = None
    unit_price: Optional[int] = None
    is_active: Optional[bool] = None
    sort_order: Optional[int] = None


@router.post("/products")
def create_product(body: ProductBody, db: Session = Depends(get_db), current_user: User = Depends(_require_manager)):
    if not body.name.strip():
        raise HTTPException(status_code=400, detail="제품명을 입력해주세요.")
    p = EnteralProduct(
        name=body.name.strip(), brand=body.brand, unit=(body.unit or "팩"),
        spec=body.spec, memo=body.memo, unit_price=body.unit_price,
        is_active=bool(body.is_active) if body.is_active is not None else True,
        sort_order=body.sort_order or 0,
    )
    db.add(p)
    db.commit()
    db.refresh(p)
    return ApiResponse(success=True, data=_product_view(p, 0))


@router.patch("/products/{pid}")
def update_product(pid: str, body: ProductBody, db: Session = Depends(get_db), current_user: User = Depends(_require_manager)):
    p = db.query(EnteralProduct).filter(EnteralProduct.id == pid).first()
    if not p:
        raise HTTPException(status_code=404, detail="제품을 찾을 수 없습니다.")
    if body.name is not None and body.name.strip():
        p.name = body.name.strip()
    if body.brand is not None:
        p.brand = body.brand
    if body.unit is not None:
        p.unit = body.unit
    if body.spec is not None:
        p.spec = body.spec
    if body.memo is not None:
        p.memo = body.memo
    if body.unit_price is not None:
        p.unit_price = body.unit_price
    if body.is_active is not None:
        p.is_active = bool(body.is_active)
    if body.sort_order is not None:
        p.sort_order = body.sort_order
    p.updated_at = now_kst()
    db.commit()
    db.refresh(p)
    return ApiResponse(success=True, data=_product_view(p, _stock_map(db).get(p.id, 0)))


@router.delete("/products/{pid}")
def delete_product(pid: str, db: Session = Depends(get_db), current_user: User = Depends(_require_manager)):
    p = db.query(EnteralProduct).filter(EnteralProduct.id == pid).first()
    if not p:
        raise HTTPException(status_code=404, detail="제품을 찾을 수 없습니다.")
    db.delete(p)
    db.commit()
    return ApiResponse(success=True, message="삭제되었습니다.")


# --------------------------------------------------------------------------- #
# 입출고 내역
# --------------------------------------------------------------------------- #
@router.get("/transactions")
def list_transactions(
    tx_type: Optional[str] = Query(None),
    product_id: Optional[str] = Query(None),
    resident: Optional[str] = Query(None),
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(_require_manager),
):
    q = db.query(EnteralTransaction)
    if tx_type in TX_TYPES:
        q = q.filter(EnteralTransaction.tx_type == tx_type)
    if product_id:
        q = q.filter(EnteralTransaction.product_id == product_id)
    if resident:
        q = q.filter(EnteralTransaction.resident_name.ilike(f"%{resident.strip()}%"))
    if start_date:
        q = q.filter(EnteralTransaction.tx_date >= start_date)
    if end_date:
        q = q.filter(EnteralTransaction.tx_date <= end_date)
    rows = q.order_by(EnteralTransaction.tx_date.desc(), EnteralTransaction.created_at.desc()).limit(5000).all()
    in_sum = sum(t.quantity for t in rows if t.tx_type == "in")
    out_sum = sum(t.quantity for t in rows if t.tx_type == "out")
    in_amt = sum((t.unit_price or 0) * t.quantity for t in rows if t.tx_type == "in")
    out_amt = sum((t.unit_price or 0) * t.quantity for t in rows if t.tx_type == "out")
    summary = {"in": in_sum, "out": out_sum, "in_amount": in_amt, "out_amount": out_amt, "count": len(rows)}
    return ApiResponse(success=True, data={"items": [_tx_view(t) for t in rows], "summary": summary})


class TxBody(BaseModel):
    product_id: Optional[str] = None
    product_name: Optional[str] = None
    tx_type: str
    quantity: int
    unit_price: Optional[int] = None
    resident_name: Optional[str] = None
    resident_id: Optional[str] = None
    tx_date: Optional[str] = None
    note: Optional[str] = None


@router.post("/transactions")
def create_transaction(body: TxBody, db: Session = Depends(get_db), current_user: User = Depends(_require_manager)):
    if body.tx_type not in TX_TYPES:
        raise HTTPException(status_code=400, detail="입고/출고 구분이 올바르지 않습니다.")
    if not body.quantity or body.quantity <= 0:
        raise HTTPException(status_code=400, detail="수량을 1 이상으로 입력해주세요.")

    name = (body.product_name or "").strip()
    p = None
    if body.product_id:
        p = db.query(EnteralProduct).filter(EnteralProduct.id == body.product_id).first()
        if not p:
            raise HTTPException(status_code=404, detail="제품을 찾을 수 없습니다.")
        name = p.name
    if not name:
        raise HTTPException(status_code=400, detail="제품을 선택해주세요.")

    # 단가: 입력값 우선, 없으면 제품 기본 단가 사용(스냅샷)
    unit_price = body.unit_price if body.unit_price is not None else (p.unit_price if p else None)
    # 입고에 단가를 명시하면 제품 기본 단가를 최신 매입가로 갱신
    if body.tx_type == "in" and body.unit_price is not None and p is not None:
        p.unit_price = body.unit_price

    t = EnteralTransaction(
        product_id=body.product_id,
        product_name=name,
        tx_type=body.tx_type,
        quantity=int(body.quantity),
        unit_price=unit_price,
        resident_name=(body.resident_name or "").strip() or None,
        resident_id=body.resident_id,
        tx_date=(body.tx_date or now_kst().strftime("%Y-%m-%d")),
        note=(body.note or "").strip() or None,
        created_by=getattr(current_user, "name", None),
    )
    db.add(t)
    db.commit()
    db.refresh(t)
    return ApiResponse(success=True, data=_tx_view(t))


@router.delete("/transactions/{tid}")
def delete_transaction(tid: str, db: Session = Depends(get_db), current_user: User = Depends(_require_manager)):
    t = db.query(EnteralTransaction).filter(EnteralTransaction.id == tid).first()
    if not t:
        raise HTTPException(status_code=404, detail="내역을 찾을 수 없습니다.")
    db.delete(t)
    db.commit()
    return ApiResponse(success=True, message="삭제되었습니다.")


# --------------------------------------------------------------------------- #
# 어르신별 비용 (출고 기준)
# --------------------------------------------------------------------------- #
@router.get("/resident-costs")
def resident_costs(
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(_require_manager),
):
    q = db.query(EnteralTransaction).filter(EnteralTransaction.tx_type == "out")
    if start_date:
        q = q.filter(EnteralTransaction.tx_date >= start_date)
    if end_date:
        q = q.filter(EnteralTransaction.tx_date <= end_date)
    rows = q.all()

    agg: dict = {}
    for t in rows:
        name = (t.resident_name or "").strip() or "(미지정)"
        amt = (t.unit_price or 0) * (t.quantity or 0)
        g = agg.setdefault(name, {"resident_name": name, "qty": 0, "amount": 0, "by_product": {}})
        g["qty"] += t.quantity
        g["amount"] += amt
        bp = g["by_product"].setdefault(t.product_name, {"product_name": t.product_name, "qty": 0, "amount": 0})
        bp["qty"] += t.quantity
        bp["amount"] += amt

    items = []
    for g in agg.values():
        items.append({
            "resident_name": g["resident_name"],
            "qty": g["qty"],
            "amount": g["amount"],
            "products": sorted(g["by_product"].values(), key=lambda x: x["amount"], reverse=True),
        })
    items.sort(key=lambda x: x["amount"], reverse=True)
    total = sum(x["amount"] for x in items)
    return ApiResponse(success=True, data={"items": items, "total": total, "count": len(rows)})


# --------------------------------------------------------------------------- #
# 입소자(어르신) 목록 — 출고 대상 선택용
# --------------------------------------------------------------------------- #
@router.get("/residents")
def list_residents(db: Session = Depends(get_db), current_user: User = Depends(_require_manager)):
    rows = (
        db.query(CareforResident)
        .filter(CareforResident.status == "active")
        .order_by(CareforResident.name.asc())
        .all()
    )
    return ApiResponse(success=True, data=[
        {"id": r.id, "name": r.name, "room_name": r.room_name} for r in rows
    ])


# --------------------------------------------------------------------------- #
# 엑셀 내보내기
# --------------------------------------------------------------------------- #
def _xlsx_response(sheets: dict, filename: str) -> StreamingResponse:
    """sheets: {sheet_name: (headers:list, rows:list[list])} → xlsx StreamingResponse"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    wb.remove(wb.active)
    head_fill = PatternFill("solid", fgColor="FFF1E6")
    head_font = Font(bold=True)
    for name, (headers, rows) in sheets.items():
        ws = wb.create_sheet(title=name[:31])
        ws.append(headers)
        for c in ws[1]:
            c.font = head_font
            c.fill = head_fill
            c.alignment = Alignment(horizontal="center")
        for r in rows:
            ws.append(r)
        # 열 너비 자동(대략)
        for i, h in enumerate(headers, start=1):
            maxlen = len(str(h))
            for r in rows:
                if i - 1 < len(r) and r[i - 1] is not None:
                    maxlen = max(maxlen, len(str(r[i - 1])))
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = min(max(maxlen + 2, 8), 40)
        ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


_TXLABEL = {"in": "입고", "out": "출고/반출"}


@router.get("/export/transactions")
def export_transactions(
    tx_type: Optional[str] = Query(None),
    product_id: Optional[str] = Query(None),
    resident: Optional[str] = Query(None),
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(_require_manager),
):
    q = db.query(EnteralTransaction)
    if tx_type in TX_TYPES:
        q = q.filter(EnteralTransaction.tx_type == tx_type)
    if product_id:
        q = q.filter(EnteralTransaction.product_id == product_id)
    if resident:
        q = q.filter(EnteralTransaction.resident_name.ilike(f"%{resident.strip()}%"))
    if start_date:
        q = q.filter(EnteralTransaction.tx_date >= start_date)
    if end_date:
        q = q.filter(EnteralTransaction.tx_date <= end_date)
    rows = q.order_by(EnteralTransaction.tx_date.desc(), EnteralTransaction.created_at.desc()).all()

    headers = ["거래일", "구분", "제품", "수량", "단가", "금액", "어르신", "메모", "작성자", "등록시각"]
    data = [[
        t.tx_date, _TXLABEL.get(t.tx_type, t.tx_type), t.product_name, t.quantity,
        t.unit_price or 0, (t.unit_price or 0) * (t.quantity or 0),
        t.resident_name or "", t.note or "", t.created_by or "",
        t.created_at.strftime("%Y-%m-%d %H:%M") if t.created_at else "",
    ] for t in rows]
    fn = f"enteral_transactions_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return _xlsx_response({"입출고내역": (headers, data)}, fn)


@router.get("/export/resident-costs")
def export_resident_costs(
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(_require_manager),
):
    q = db.query(EnteralTransaction).filter(EnteralTransaction.tx_type == "out")
    if start_date:
        q = q.filter(EnteralTransaction.tx_date >= start_date)
    if end_date:
        q = q.filter(EnteralTransaction.tx_date <= end_date)
    rows = q.all()

    agg: dict = {}
    detail = []
    for t in rows:
        name = (t.resident_name or "").strip() or "(미지정)"
        amt = (t.unit_price or 0) * (t.quantity or 0)
        g = agg.setdefault(name, {"qty": 0, "amount": 0})
        g["qty"] += t.quantity
        g["amount"] += amt
        detail.append([name, t.product_name, t.tx_date, t.quantity, t.unit_price or 0, amt])

    summary = sorted(([n, g["qty"], g["amount"]] for n, g in agg.items()), key=lambda x: x[2], reverse=True)
    detail.sort(key=lambda x: (x[0], x[2]))
    period = f"{start_date or '전체'}~{end_date or ''}".strip("~")
    sheets = {
        "어르신별 합계": (["어르신", "출고수량", "비용(원)"], summary),
        "제품별 상세": (["어르신", "제품", "거래일", "수량", "단가", "금액"], detail),
    }
    fn = f"enteral_resident_costs_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return _xlsx_response(sheets, fn)


@router.get("/export/stock")
def export_stock(db: Session = Depends(get_db), current_user: User = Depends(_require_manager)):
    stock = _stock_map(db)
    rows = db.query(EnteralProduct).order_by(EnteralProduct.sort_order.asc(), EnteralProduct.name.asc()).all()
    headers = ["제품명", "브랜드", "규격", "단위", "단가", "현재고", "활성"]
    data = [[p.name, p.brand or "", p.spec or "", p.unit or "", p.unit_price or 0, stock.get(p.id, 0), "Y" if p.is_active else "N"] for p in rows]
    fn = f"enteral_stock_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return _xlsx_response({"재고현황": (headers, data)}, fn)
